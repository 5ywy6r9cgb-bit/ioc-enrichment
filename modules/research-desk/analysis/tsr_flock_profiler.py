#!/usr/bin/env python3
"""
TSR Flock Dataset Profiler — streaming summary stats for large Flock
spreadsheet exports (xlsx/csv), without loading whole files into memory.

Designed for the 8GB-RAM MacBook Air constraint: uses openpyxl in
read_only + iter_rows streaming mode for xlsx, and the csv module
(line-by-line) for csv/tsv. Never holds a full sheet in memory.

For each file, reports:
  - row / column counts
  - detected date column(s) and their min/max range
  - detected agency column and top values by frequency
  - detected incident-code / reason column and top values by frequency
  - detected "response type" column (lookup/search/hotlist hit, etc.)

Usage:
    python3 tsr_flock_profiler.py --dir ~/Downloads/flock_columbus_export \\
        --out flock_profile_summary.json

Requires: openpyxl (pip install openpyxl --break-system-packages)
"""

import argparse
import csv
import json
import os
import sys
from collections import Counter
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

SPREADSHEET_EXT = {".xlsx", ".xlsm"}
DELIMITED_EXT = {".csv", ".tsv"}

# Column name fragments used to guess semantic role. Case-insensitive substring match.
AGENCY_HINTS = ["agency", "organization", "org name", "department"]
DATE_HINTS = ["date", "timestamp", "utc", "time"]
CODE_HINTS = ["incident", "reason", "code", "hotlist", "hit type"]
RESPONSE_TYPE_HINTS = ["response", "type", "lookup", "search"]

TOP_N = 25


def guess_role(header_name):
    h = header_name.lower().strip()
    if any(frag in h for frag in AGENCY_HINTS):
        return "agency"
    if any(frag in h for frag in DATE_HINTS):
        return "date"
    if any(frag in h for frag in CODE_HINTS):
        return "code"
    if any(frag in h for frag in RESPONSE_TYPE_HINTS):
        return "response_type"
    return None


def try_parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def profile_rows(header, row_iterator, filename):
    roles = {i: guess_role(h) for i, h in enumerate(header) if h}
    agency_col = next((i for i, r in roles.items() if r == "agency"), None)
    date_cols = [i for i, r in roles.items() if r == "date"]
    code_col = next((i for i, r in roles.items() if r == "code"), None)
    resp_col = next((i for i, r in roles.items() if r == "response_type"), None)

    row_count = 0
    agency_counter = Counter()
    code_counter = Counter()
    resp_counter = Counter()
    date_min, date_max = None, None
    bad_date_samples = 0

    for row in row_iterator:
        row_count += 1
        if agency_col is not None and agency_col < len(row) and row[agency_col]:
            agency_counter[str(row[agency_col]).strip()] += 1
        if code_col is not None and code_col < len(row) and row[code_col]:
            code_counter[str(row[code_col]).strip()] += 1
        if resp_col is not None and resp_col < len(row) and row[resp_col]:
            resp_counter[str(row[resp_col]).strip()] += 1
        for dc in date_cols:
            if dc < len(row) and row[dc]:
                dt = try_parse_date(row[dc])
                if dt:
                    if date_min is None or dt < date_min:
                        date_min = dt
                    if date_max is None or dt > date_max:
                        date_max = dt
                else:
                    bad_date_samples += 1
        if row_count % 250000 == 0:
            print(f"    ...{row_count:,} rows scanned in {filename}")

    return {
        "row_count": row_count,
        "column_count": len(header),
        "columns": header,
        "detected_roles": {header[i]: r for i, r in roles.items() if r},
        "date_range": {
            "min": date_min.isoformat() if date_min else None,
            "max": date_max.isoformat() if date_max else None,
            "unparsed_date_values_sampled": bad_date_samples,
        } if date_cols else None,
        "top_agencies": agency_counter.most_common(TOP_N) if agency_col is not None else None,
        "distinct_agency_count": len(agency_counter) if agency_col is not None else None,
        "top_incident_codes": code_counter.most_common(TOP_N) if code_col is not None else None,
        "top_response_types": resp_counter.most_common(TOP_N) if resp_col is not None else None,
    }


def profile_xlsx(path):
    if load_workbook is None:
        return {"error": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"}
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        row_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(row_iter)]
        except StopIteration:
            sheets[ws.title] = {"row_count": 0, "column_count": 0, "columns": []}
            continue
        sheets[ws.title] = profile_rows(header, row_iter, os.path.basename(path))
    wb.close()
    return {"sheets": sheets}


def profile_delimited(path, delimiter):
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            return {"row_count": 0, "column_count": 0, "columns": []}
        return profile_rows(header, reader, os.path.basename(path))


def main():
    ap = argparse.ArgumentParser(description="Profile a directory of large Flock spreadsheet exports")
    ap.add_argument("--dir", required=True, help="Directory containing xlsx/csv files")
    ap.add_argument("--out", required=True, help="Output summary JSON path")
    args = ap.parse_args()

    root = os.path.abspath(args.dir)
    if not os.path.isdir(root):
        print(f"error: not a directory: {root}", file=sys.stderr)
        sys.exit(1)

    results = {}
    for dirpath, _dirs, files in os.walk(root):
        for fn in sorted(files):
            if fn.startswith("."):
                continue
            ext = os.path.splitext(fn)[1].lower()
            full_path = os.path.join(dirpath, fn)
            rel_path = os.path.relpath(full_path, root)

            if ext in SPREADSHEET_EXT:
                print(f"Profiling (xlsx): {rel_path}")
                results[rel_path] = profile_xlsx(full_path)
            elif ext in DELIMITED_EXT:
                delim = "\t" if ext == ".tsv" else ","
                print(f"Profiling (csv):  {rel_path}")
                results[rel_path] = profile_delimited(full_path, delim)
            else:
                continue

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    print(f"\nProfile summary written: {args.out}")
    print(f"Files profiled: {len(results)}")


if __name__ == "__main__":
    main()
